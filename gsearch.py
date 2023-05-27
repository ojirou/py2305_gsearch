import requests
import bs4
import pandas as pd
import subprocess
import os
import glob
import openpyxl as px
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
def main():
    Dir=r'C:\\Users\\user\\python\\web_scraping\\GoogleSearch\\'
    search_keyword= input('キーワード　>> ') 
    search_keyword2=search_keyword.replace(" ", "_")
    ExcelName=Dir+search_keyword2+'.xlsx'
    search_url = 'https://www.google.co.jp/search?hl=ja&num=100&q=' + search_keyword
    res_google = requests.get(search_url)
    res_google.raise_for_status()
    bs4_google = bs4.BeautifulSoup(res_google.text, 'lxml')
    google_search_page = bs4_google.select('div.kCrYT>a')
    rank = 1
    site_rank = []
    site_title = []
    site_url = []
    for site in google_search_page:
        try:    
            site_title.append(site.select('h3.zBAuLc')[0].text)
            site_url.append(site.get('href').split('&sa=U&')[0].replace('/url?q=', ''))
            site_rank.append(rank)
            rank +=1
        except IndexError:
            continue
    df = pd.DataFrame({'順位':site_rank, 'タイトル':site_title, 'URL':site_url})
    df.to_excel(ExcelName, index=False)
    FileName=ExcelName
    name, ext=os.path.splitext(FileName)
    wb = load_workbook(FileName)
    ws = wb.active
    column = ws['C']
    for cell in column:
        cell.value = str(cell.value)
        # cell.value = "'" + str(cell.value)
    ws.column_dimensions['B'].width=60
    ws.column_dimensions['C'].width=100
    column_num=3
    row_nums=len(df)
    for row_num in range(1, row_nums+2):
        target_cell = ws.cell(row=row_num, column=column_num) 
        Address=target_cell.value
        target_cell.value = '=HYPERLINK("'+Address+'", "'+Address+'")' 
        target_cell.font  = Font(size=9, color=Color(rgb=None, indexed=None, auto=None, theme=10, tint=0.0, type="theme"))
    FileNameR=name+'_r.xlsx'
    wb.save(FileNameR)
#    os.remove(FileNameR)
    subprocess.Popen(['start', FileNameR], shell=True)
    os.startfile(Dir, operation='open')
if __name__ == "__main__":
    main()